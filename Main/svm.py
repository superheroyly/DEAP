# Using SVM

import os
import random
import numpy as np
import time
import torch
from sklearn import svm,metrics
import generate_networks
import argparse
from sklearn.preprocessing import normalize
import openpyxl
import send_mail as sm
from sklearn.model_selection import KFold
from sklearn import preprocessing
from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV, cross_val_score


path = "D:\\Pycharm\\Project\\ML\\Dataset\\DEAP\\SI6s\\"  # 数据集路径

# 读入数据和标签
with open(path + 'data.npy', 'rb') as fileTrain:
    data = np.load(fileTrain)
with open(path + 'label.npy', 'rb') as fileTrainL:
    label_train = np.load(fileTrainL)

label = label_train[:, :1]  # 取第二列的标签
label = np.ravel(label)

for i in range(len(label)):
    if label[i] > 5:
        label[i] = 1
    else:
        label[i] = 0

index = [i for i in range(len(data))]
seed=42
print("Data OK!")


def save(row, col, result):
    '''
    row, col: The first grid (row, col)
    result: Save object
    '''
    xlsx_path = 'Result/AAAI Data/4.xlsx'  # xlsx路径
    data = openpyxl.load_workbook(xlsx_path)  # open file
    sheetnames = data.sheetnames
    table = data[sheetnames[0]]  # 指定xlsx表中的sheet
    table = data.active  # Open the first sheet
    table.cell(row, col).value = result  # Set(row, col) = result
    data.save(xlsx_path)


if __name__ == '__main__':

    row, col = 4, 2  # 保存所有指标

    parameters = [{'C':[1, 3, 5, 7, 9],
                  'gamma':[0.0001, 0.001, 0.01, 0.1, 1, 10, 100],
                  'kernel':['rbf']},
                 {'C': [0.5, 1, 3, 5],
                  'kernel':['linear']}]

    svc = SVC()

    # clf = GridSearchCV(svc, parameters, cv=5)  # 网格搜索，寻找最优参数
    # clf.fit(data, label)
    # print(clf.best_params_)

    clf = SVC(kernel='rbf', C=5, gamma=1)  # 五折交叉验证
    # {'C': 5, 'gamma': 1, 'kernel': 'rbf'}
    acc = cross_val_score(clf, data, label, cv=5)
    precision = cross_val_score(clf, data, label, cv=5, scoring="precision_weighted")
    f1 = cross_val_score(clf, data, label, cv=5, scoring="f1_weighted")

    print("Acc:{0}, Precision:{1}, F1:{2}".format(round(np.mean(acc), 4), round(np.mean(precision), 4), round(np.mean(f1), 4)))

