# Leaky integrator model of Echo State Network

import numpy as np
import time
import openpyxl
import generate_networks
import random
from sklearn.preprocessing import normalize
import send_mail as sm
import joblib
import os


# 划分训练集，验证集，测试集
path = "D:\\Pycharm\\Project\\ML\\Dataset\\DEAP\\data_preprocessed_python\\DEAP_signal\\"  # 32W 8W 每个subject按4：1划分

# Train set
with open(path + 'data_training.npy', 'rb') as fileTrain:
    data_train = np.load(fileTrain)
with open(path + 'label_training.npy', 'rb') as fileTrainL:
    label_train = np.load(fileTrainL)
data_train = normalize(data_train)
# data_train = data_train[:320000, :]
# label_train = np.ravel(label_train[:, :2])  # 多维转一维 标签：Valence Arousal Domain Like
label_train = label_train[:, :2]

# Test set
with open(path + 'data_testing.npy', 'rb') as fileTest:
    data_test = np.load(fileTest)
with open(path + 'label_testing.npy', 'rb') as fileTestL:
    label_test = np.load(fileTestL)
data_test = normalize(data_test)
# data_test = data_test[:80000, :]
# label_test = np.ravel(label_test[:, [1]])
# label_test = np.ravel(label_test[:, :2])
label_test = label_test[:, :2]

len_train = len(data_train)
len_test = len(data_test)


print('Trainset and Testset OK!')


def correct_dimensions(s, targetlength):
    if s is not None:
        s = np.array(s)
        if s.ndim == 0:
            s = np.array([s] * targetlength)
        elif s.ndim == 1:
            if not len(s) == targetlength:
                raise ValueError("arg must have length " + str(targetlength))
        else:
            raise ValueError("Invalid argument")
    return s


def identity(x):
    return x


def step_function(x):
    if x > 0.5:
        return 1
    else:
        return 0


def sigmoid(x):
    return 1 / (1 + np.exp(-10 * x + 1))


def memory_capacity(L, buffer, data, output_data):
    MC = 0
    for k in range(L):
        cov_matrix = np.cov(
            np.array([data[len_train + buffer - (k + 1): len_train + buffer - (k + 1) + 1000], output_data.T[k]]))
        MC_k = cov_matrix[0][1] ** 2
        MC_k = MC_k / (np.var(data[len_train + buffer:]) * np.var(output_data.T[k]))
        MC += MC_k
    print("MC:", MC)
    return MC


class LI_ESN_internal:
    def __init__(self, n_inputs, n_outputs, n_reservoir=200, W=None, W_in=None,
                 noise=0.001, input_shift=None,
                 input_scaling=None, feedback_scaling=None,
                 teacher_scaling=None, teacher_shift=None,
                 out_activation=identity, inverse_out_activation=identity,
                 random_state=None, time_scale=None):
        # check for proper dimensionality of all arguments and write them down.
        self.n_inputs = n_inputs
        self.n_reservoir = n_reservoir
        self.n_outputs = n_outputs
        self.noise = noise
        self.input_shift = correct_dimensions(input_shift, n_inputs)
        self.input_scaling = correct_dimensions(input_scaling, n_inputs)

        self.teacher_scaling = teacher_scaling
        self.teacher_shift = teacher_shift

        self.out_activation = out_activation
        self.inverse_out_activation = inverse_out_activation
        self.random_state = random_state
        self.time_scale = time_scale
        self.W = W
        self.W_in = W_in

        # the given random_state might be either an actual RandomState object,
        # a seed or None (in which case we use numpy's builtin RandomState)
        if isinstance(random_state, np.random.RandomState):
            self.random_state_ = random_state
        elif random_state:
            try:
                self.random_state_ = np.random.RandomState(random_state)
            except TypeError as e:
                raise Exception("Invalid seed: " + str(e))
        else:
            self.random_state_ = np.random.mtrand._rand

    def _scale_inputs(self, inputs):
        """for each input dimension j: multiplies by the j'th entry in the
        input_scaling argument, then adds the j'th entry of the input_shift
        argument."""
        if self.input_scaling is not None:
            inputs = np.dot(inputs, np.diag(self.input_scaling))
        if self.input_shift is not None:
            inputs = inputs + self.input_shift
        return inputs

    def _update(self, state, input_pattern):
        # leaky integrator model:
        # it can adjust timescales for each neurons.
        preactivation = (np.dot(self.W, state) + np.dot(self.W_in, input_pattern))
        state = (1 - self.time_scale) * state + self.time_scale * np.tanh(preactivation)
        # state = (1 - self.time_scale) * state + self.time_scale * sigmoid(preactivation)

        t2 = state + self.noise * self.time_scale * (self.random_state_.rand(self.n_reservoir) - 0.5)
        return t2

    def calc_lyapunov_exp(self, inputs, initial_distance, n):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        states1 = np.zeros((inputs.shape[0], self.n_reservoir))
        states2 = np.zeros((inputs.shape[0], self.n_reservoir))
        transient = min(int(inputs.shape[0] / 10), 100)
        for i in range(1, transient):
            states1[i, :] = self._update(states1[i - 1], inputs[i, :])
        states2[transient - 1, :] = states1[transient - 1, :]
        states2[transient - 1, n] = states2[transient - 1, n] + initial_distance
        gamma_k_list = []
        for k in range(transient, inputs.shape[0]):
            states1[k, :] = self._update(states1[k - 1], inputs[k, :])
            states2[k, :] = self._update(states2[k - 1], inputs[k, :])
            gamma_k = np.linalg.norm(states2[k, :] - states1[k, :])
            gamma_k_list.append(gamma_k / initial_distance)
            states2[k, :] = states1[k, :] + (initial_distance / gamma_k) * (states2[k, :] - states1[k, :])
        lyapunov_exp = np.mean(np.log(gamma_k_list))
        return lyapunov_exp

    def fit(self, inputs, outputs):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        if outputs.ndim < 2:
            outputs = np.reshape(outputs, (len(outputs), -1)) # , 1

        inputs_scaled = self._scale_inputs(inputs)
        teachers_scaled = outputs

        # step the reservoir through the given input,output pairs:
        # states = np.zeros((inputs.shape[0], self.n_reservoir))
        states = np.random.normal(0, 0.1**0.5, (inputs.shape[0], self.n_reservoir))  # 设置状态为均值为0，方差为0.1的正态分布，使参数更容易收敛。
        # states = joblib.load('./model/AAAI model/compare/state.pkl')
        self.st = states
        for n in range(1, inputs.shape[0]):
            temp = self._update(states[n - 1], inputs_scaled[n, :])
            states[n, :] = temp
        transient = min(int(inputs.shape[0] / 10), 100)
        extended_states = np.hstack((states, inputs_scaled))

        self.W_out = np.dot(np.linalg.pinv(extended_states[transient:, :]), teachers_scaled[transient:, :]).T
        # print(self.W_out.shape)

        # remember the last state for later:
        self.laststate = states[-1, :]
        self.lastinput = inputs[-1, :]
        self.lastoutput = teachers_scaled[-1, :]



        # apply learned weights to the collected states:
        pred_train = np.dot(extended_states, self.W_out.T)
        return pred_train

    def predict(self, inputs, continuation=True, use_model=False):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        n_samples = inputs.shape[0]

        if continuation:
            if use_model == False:
                laststate = self.laststate
                lastinput = self.lastinput
                lastoutput = self.lastoutput
                W_out = self.W_out
            else:
                laststate = joblib.load('./model/Test/laststate.pkl')
                lastinput = joblib.load('./model/Test/lastinput.pkl')
                lastoutput = joblib.load('./model/Test/lastoutput.pkl')
                W_out = joblib.load('./model/Test/W_out.pkl')
        else:
            laststate = np.zeros(self.n_reservoir)
            lastinput = np.zeros(self.n_inputs)
            lastoutput = np.zeros(self.n_outputs)

        inputs = inputs
        states = np.vstack(
            [laststate, np.zeros((n_samples, self.n_reservoir))])
        outputs = np.vstack(
            [lastoutput, np.zeros((n_samples, self.n_outputs))])
        # print(inputs.shape) # (78080, 70)
        # print(states.shape) # (78081, 200)
        for n in range(n_samples):
            states[n + 1, :] = self._update(states[n, :], inputs[n, :])
            t1 = np.concatenate([states[n + 1, :], inputs[n, :]])
            outputs[n + 1, :] = np.dot(W_out, t1)

        if use_model == True:
            return self.out_activation(outputs[1:])
        # print("outputs.shape:", outputs.shape)
        weight = [self.laststate, self.lastinput, self.lastoutput, self.W_out, self.st]
        return self.out_activation(outputs[1:]), weight
        # print(outputs[1:])
        # return np.heaviside(outputs[1:]-0.5, 0)*0.3


def make_data_for_narma(length):
    tau = 0.01
    buffer = 100
    x = np.random.rand(length + 100) * 0.5
    y = np.zeros(length)
    for i in range(length):
        if i < 29:
            y[i] = 0.2 * y[i - 1] + 0.004 * y[i - 1] * np.sum(np.hstack((y[i - 29:], y[:i]))) + 1.5 * x[i - 29 + 100] * \
                   x[i + 100] + 0.001
        else:
            y[i] = 0.2 * y[i - 1] + 0.004 * y[i - 1] * np.sum(np.hstack((y[i - 29:i]))) + 1.5 * x[i - 29 + 100] * x[
                i + 100] + 0.001
    return x, y


def computeedge(g):
    all = 0
    for i in range(g.shape[0]):
        for j in range(g.shape[1]):
            if g[i][j] != 0:
                all += 1
    return all



def calculate_narma(p1, p2, n, average_degree, m, is_layered, use_model):
    '''
    :param p1: 模块内连接概率
    :param p2: 模块间连接概率
    :param n: 储层节点个数
    :param average_degree: 平均度
    :param m: 模块数量
    :param is_layered: False=模块化
    :return: resulturacy  F1-score
    '''

    global Acc_max
    # path = './model/AAAI model/p1_fixed p2 m n(p1=0.1 n=600)'
    path = './model/AAAI model/compare'
    for k in range(1):  # 重复实验次数
        degree, Wtmp = [], []

        if use_model == False:
            if not os.path.exists(path):
                os.mkdir(path)

            # W_IN
            W_IN = np.random.uniform(-0.1, 0.1, size=(n, 128))  # size的第二参数表示输入维度

            # W
            if is_layered == True:  # Deep-ESN
                W = generate_networks.make_recurrent_layered_network(n, average_degree, m, p1)
            elif is_layered == False and m != 1:  # Modular-ESN
                W, degree = generate_networks.make_modular_network(n, average_degree, m, p1, p2)
                Wtmp = W
                # pass
            elif m == 1:  # 传统ESN
                W = generate_networks.make_ESN(n, p1)
            # W = joblib.load(path + '/W.pkl')
            # W_IN = joblib.load(path + '/W_IN.pkl')
            # 调整谱半径
            radius = np.max(np.abs(np.linalg.eigvals(W)))
            W = W * (SPECT_RADIUS / radius)
            radius = np.max(np.abs(np.linalg.eigvals(W)))
            print("spectral_radius:", round(radius, 4))

            # save_point(4,9,degree)  # 保存节点度
        else:
            W_IN = joblib.load(path+'/W_IN.pkl')
            W = joblib.load(path+'/W.pkl')


        # 构建泄露积分型模块化ESN
        esn = LI_ESN_internal(n_inputs=128,
                              n_outputs=2,
                              n_reservoir=n,
                              W=W,
                              W_in=W_IN,
                              noise=0.000,
                              time_scale=time_scale)

        if use_model == False:
            # 拟合训练集
            esn.fit(data_train, label_train)
            # 预测测试集
            prediction, weight = esn.predict(data_test, continuation=True, use_model=use_model)
        else:
            prediction = esn.predict(data_test, continuation=True, use_model=use_model)

        # tp, tn, fp, fn = 0, 0, 0, 0
        #
        # for i in range(label_test):
        #     # 分类任务
        #     if (prediction[i] > 5 and label_test[i] > 5):
        #         tp += 1
        #     elif (prediction[i] <= 5 and label_test[i] <= 5):
        #         tn += 1
        #     elif (prediction[i] > 5 and label_test[i] <= 5):
        #         fp += 1
        #     else:
        #         fn += 1
        k, yes, all_num = 0, 0, 0
        for i in range(len_test):
            if label_test[i][1] > 5 and label_test[i][0] > 5:  # stress !
                k = k + (prediction[i][0] - label_test[i][0]) ** 2 + (prediction[i][1] - label_test[i][1]) ** 2  # square difference
                all_num += 1
                if prediction[i][1] > 5 and prediction[i][0] > 5:
                    # print(prediction[i])
                    yes += 1
            if label_test[i][1] <= 5 and label_test[i][0] <= 5:  # stress !
                k = k + (prediction[i][0] - label_test[i][0]) ** 2 + (prediction[i][1] - label_test[i][1]) ** 2  # square difference
                all_num += 1
                if prediction[i][1] <= 5 and prediction[i][0] <= 5:
                    yes += 1
            if label_test[i][1] > 5 and label_test[i][0] <= 5:  # stress !
                k = k + (prediction[i][0] - label_test[i][0]) ** 2 + (prediction[i][1] - label_test[i][1]) ** 2  # square difference
                all_num += 1
                if prediction[i][1] > 5 and prediction[i][0] <= 5:
                    yes += 1
            if label_test[i][1] <= 5 and label_test[i][0] > 5:  # stress !
                k = k + (prediction[i][0] - label_test[i][0]) ** 2 + (prediction[i][1] - label_test[i][1]) ** 2  # square difference
                all_num += 1
                if prediction[i][1] <= 5 and prediction[i][0] > 5:
                    yes += 1
        result = yes / all_num
        print(yes, all_num, result)
        # result = [tp, tn, fp, fn]
        # result.append(round(result[0]/(result[0]+result[2]), 6))  # precision
        # result.append(round(result[0]/(result[0]+result[3]), 6))  # recall
        # result.append(round(result[0]/(result[0]+result[3]), 6))  # sensitivity
        # result.append(round(result[2]/(result[2]+result[3]), 6))  # specificity
        # result.append(round((result[0]+result[1]) / (sum(result[0:4])), 6))  # accuracy
        # result.append(round(2*result[5]*result[4] / (result[5]+result[4]), 6))  # f1-score
        # print(result)

        # save model weight
        # if use_model == False and result[8] > Acc_max:
        #     Acc_max = result[8]
        #     joblib.dump(Wtmp, path+'/W.pkl')
        #     joblib.dump(W_IN, path+'/W_IN.pkl')
        #     joblib.dump(weight[0], path+'/laststate.pkl')
        #     joblib.dump(weight[1], path+'/lastinput.pkl')
        #     joblib.dump(weight[2], path+'/lastoutput.pkl')
        #     joblib.dump(weight[3], path+'/W_out.pkl')
        #     joblib.dump(weight[4], path+'/state.pkl')


    return result

def save_point(row, col, ac):
    # 保存节点入度，出度，分布
    xlsx_path = 'Result/Data/degree.xlsx'
    data = openpyxl.load_workbook(xlsx_path)
    sheetnames = data.sheetnames
    table = data[sheetnames[0]] # 指定xlsx表中的sheet
    table = data.active
    for i in ac:
        table.cell(row, col).value = i
        row += 1
    data.save(xlsx_path)


def save(row, col, result):
    xlsx_path = 'Result/AAAI Data/test.xlsx'  # xlsx路径
    data = openpyxl.load_workbook(xlsx_path)
    sheetnames = data.sheetnames
    table = data[sheetnames[0]] # 指定xlsx表中的sheet
    table = data.active
    for i in result:
        table.cell(row, col).value = i
        col += 1
    data.save(xlsx_path)

def save_acc(row, col, acc):
    xlsx_path = 'Result/AAAI Data/p1_fixed p2 m n.xlsx'  # xlsx路径
    data = openpyxl.load_workbook(xlsx_path)
    sheetnames = data.sheetnames
    table = data[sheetnames[0]] # 指定xlsx表中的sheet
    table = data.active
    table.cell(row, col).value = acc
    data.save(xlsx_path)


if __name__ == '__main__':

    all_time = time.time()

    # 结果
    Result = []
    Acc_max = 1.5

    # 自变量
    m_list = np.array([1,2,3,4,5,6])  # 模块化数量 Modularity
    # m_list = np.array([1])
    p1_list = np.arange(0.20, 0.201, 0.05)  # 模块内连接概率
    p2_list = np.arange(0.02, 0.101, 0.02)  # 模块间连接概率
    n_list = np.array([720, 1440, 2160, 2880])  # 储层节点数
    r_list = np.arange(0.1, 0.11, 0.2)  # 谱半径
    l_list = np.arange(0.06, 0.061, 0.01)  # 泄露率

    row, col = 4, 2  # 保存所有指标
    # row1, col1 = 238, 16  # 保存Accuracy

    for ti in range(1):
        for n in n_list:
            for r in r_list:
                tem = []
                start_time = time.time()  # 计时单次运行时间

                N_node = n # 储层节点数
                a = 0.1  # 泄露积分型中的参数a
                time_scale = np.ones(N_node) * a
                SPECT_RADIUS = 0.1  # 谱半径
                # Result Degree
                result = calculate_narma(p1=0.05,
                                         p2=0.01,
                                         n=N_node,
                                         average_degree=10,
                                         m=6,
                                         is_layered=False,
                                         use_model=False)
                end_time = time.time() - start_time

                print('***' * 24)
                print('---' * 11 + "Result" + '---' * 11)
                print("p1={0}, p2={1}, n={2}, m={3}, r={4}, a={5}, runtime={6},"
                      .format(0.05, round(0.01, 2), N_node, 6, SPECT_RADIUS, a, round(end_time, 2)))
                # print("result = {0}".format(result))
                # print("Acc={0}, F1-score={1}".format(result[8], result[9]))
                print("Acc={0}, F1-score={1}".format(result, result))
                print('---' * 24)
                print('***' * 24)

                # tem.extend([result])
                # save(row, col, result)
                row += 1
                Result.append(result)

        row += 5


    # print("p1_list:", p1_list)
    print("p2_list:", p2_list)
    print("m_list:", m_list)
    print("Result:", Result)

    # 发送邮件
    all_time = time.time()-all_time
    print("工作站 All Runtime:{0}s".format(all_time))
    context = "工作站 Runtime:" + str(all_time) + "s \nOK!"
    sm.send_mail(context)

