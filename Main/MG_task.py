# Leaky integrator model of Echo State Network
import numpy as np
import time
import openpyxl
import generate_networks
import argparse
from scipy import linalg

# 划分训练集，验证集，测试集
# path = "D:\\Pycharm\\Project\\ML\\Dataset\\MG\\MackeyGlass_t17.txt"
path = "D:\\Pycharm\\Project\\ML\\Dataset\\MG\\MG.txt"
trainLen = 6000
testLen = 3999
initLen = 100
data = np.loadtxt(path)
para_y = []
print("Dataset OK！")

def correct_dimensions(s, targetlength):
    if s is not None:
        s = np.array(s)
        if s.ndim == 0:
            s = np.array([s] * targetlength)
        elif s.ndim == 1:
            if not len(s) == targetlength:
                raise ValueError("arg must have length " + str(targetlength))
        else:
            raise ValueError("Invalid argument")
    return s


def identity(x):
    return x


def step_function(x):
    if x > 0.5:
        return 1
    else:
        return 0


def sigmoid(x):
    return 1 / (1 + np.exp(-10 * x + 1))


class LI_ESN_internal:
    def __init__(self, n_inputs, n_outputs, n_reservoir=200, W=None, W_in=None,
                 noise=0.001, input_shift=None,
                 input_scaling=None, feedback_scaling=None,
                 teacher_scaling=None, teacher_shift=None,
                 out_activation=identity, inverse_out_activation=identity,
                 random_state=None, time_scale=None):
        # check for proper dimensionality of all arguments and write them down.
        self.n_inputs = n_inputs
        self.n_reservoir = n_reservoir
        self.n_outputs = n_outputs
        self.noise = noise
        self.input_shift = correct_dimensions(input_shift, n_inputs)
        self.input_scaling = correct_dimensions(input_scaling, n_inputs)

        self.teacher_scaling = teacher_scaling
        self.teacher_shift = teacher_shift

        self.out_activation = out_activation
        self.inverse_out_activation = inverse_out_activation
        self.random_state = random_state
        self.time_scale = time_scale
        self.W = W
        self.W_in = W_in

        # the given random_state might be either an actual RandomState object,
        # a seed or None (in which case we use numpy's builtin RandomState)
        if isinstance(random_state, np.random.RandomState):
            self.random_state_ = random_state
        elif random_state:
            try:
                self.random_state_ = np.random.RandomState(random_state)
            except TypeError as e:
                raise Exception("Invalid seed: " + str(e))
        else:
            self.random_state_ = np.random.mtrand._rand

    def _scale_inputs(self, inputs):
        """for each input dimension j: multiplies by the j'th entry in the
        input_scaling argument, then adds the j'th entry of the input_shift
        argument."""
        if self.input_scaling is not None:
            inputs = np.dot(inputs, np.diag(self.input_scaling))
        if self.input_shift is not None:
            inputs = inputs + self.input_shift
        return inputs

    def _update(self, state, input_pattern):
        # leaky integrator model:
        # it can adjust timescales for each neurons.
        preactivation = (np.dot(self.W, state) + np.dot(self.W_in, input_pattern))
        state = (1 - self.time_scale) * state + self.time_scale * np.tanh(preactivation)
        # state = (1 - self.time_scale) * state + self.time_scale * sigmoid(preactivation)

        t2 = state + self.noise * self.time_scale * (self.random_state_.rand(self.n_reservoir) - 0.5)
        return t2

    def calc_lyapunov_exp(self, inputs, initial_distance, n):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        states1 = np.zeros((inputs.shape[0], self.n_reservoir))
        states2 = np.zeros((inputs.shape[0], self.n_reservoir))
        transient = min(int(inputs.shape[0] / 10), 100)
        for i in range(1, transient):
            states1[i, :] = self._update(states1[i - 1], inputs[i, :])
        states2[transient - 1, :] = states1[transient - 1, :]
        states2[transient - 1, n] = states2[transient - 1, n] + initial_distance
        gamma_k_list = []
        for k in range(transient, inputs.shape[0]):
            states1[k, :] = self._update(states1[k - 1], inputs[k, :])
            states2[k, :] = self._update(states2[k - 1], inputs[k, :])
            gamma_k = np.linalg.norm(states2[k, :] - states1[k, :])
            gamma_k_list.append(gamma_k / initial_distance)
            states2[k, :] = states1[k, :] + (initial_distance / gamma_k) * (states2[k, :] - states1[k, :])
        lyapunov_exp = np.mean(np.log(gamma_k_list))
        return lyapunov_exp

    def fit_predict(self, trainLen):
        X = np.zeros((1+self.n_inputs+self.n_reservoir, trainLen-initLen))
        Yt = data[None, initLen+1: trainLen+1]
        x = np.zeros((self.n_reservoir, 1))  # 矩阵x[1000,1]


        for t in range(trainLen):  # 训练2000次
            u = data[t]  # 输入数据u，读入一行数据
            x = (1 - a) * x + a * np.tanh(np.dot(self.W, x) + np.dot(self.W_in, np.vstack((1, u))))

            if t >= initLen:
                tem = np.vstack((1, u, x))
                X[:, t - initLen] = np.vstack((1, u, x))[:, 0]  # [[1] [u] [...]...] 垂直堆叠-->(1002, 1)

        # train the output by ridge regression 通过岭回归训练输出
        reg = 1e-8  # regularization coefficient 正则化系数
        # direct equations from texts:
        # X_T = X.T
        # Wout = np.dot( np.dot(Yt,X_T), linalg.inv( np.dot(X,X_T) + \
        #    reg*np.eye(1+inSize+resSize) ) )
        # using scipy.linalg.solve:
        # numpy.linalg.solve(a, b)  a系数矩阵，b纵坐标或因变量的值，返回值意义：方程组ax=b的解
        # np.eye()作用是返回一个对角线上全是1，而其他位置全为0的一个二维数组
        Wout = linalg.solve(np.dot(X, X.T) + reg * np.eye(1 + self.n_inputs + self.n_reservoir),
                            np.dot(X, Yt.T)).T  # (1, 1002)

        # run the trained ESN in a generative mode. no need to initialize here,
        # because x is initialized with training data and we continue from there.
        Y = np.zeros((self.n_outputs, testLen))
        u = data[trainLen]
        Uk = np.array([])
        Yk = np.array([])
        for t in range(testLen):  # 测试数据集
            # print(u)
            Uk = np.append(Uk, u)
            # print("Uk:", Uk)
            x = (1 - a) * x + a * np.tanh(np.dot(self.W_in, np.vstack((1, u))) + np.dot(self.W, x))
            # 输出矩阵 * 此刻状态矩阵 = 此刻预测值
            y = np.dot(Wout, np.vstack((1, u, x)))  # (1, 1002) (1002, 1) (1,1)
            # t时刻的预测值 Y: 1 * testLen
            Y[:, t] = y
            Yk = np.append(Yk, y)
            # print("Yk:", Yk)
            # generative mode:
            u = y

        pre_step = 0
        # compute MSE for the first errorLen time steps 用MSE计算errorLen时间步长
        chance = 0
        errorLen = 3600
        # for t in range(errorLen):
        #     val = abs(data[trainLen+1+t]-Y[0][t]) / data[trainLen+1+t]
        #     if val < 0.10:
        #         pre_step += 1
        #     elif val >=0.10 and chance < 5:
        #         chance +=1
        #     elif chance >= 5:
        #         break
        avg_y = sum(Y[0, 0:errorLen]) / errorLen
        fc = sum(np.square(data[trainLen + 1:trainLen + 1 + errorLen] -
                           avg_y)) / errorLen
        nrmse = sum(np.square(data[trainLen + 1:trainLen + errorLen + 1] -
                              Y[0, 0:errorLen])) / errorLen
        nrmse = (nrmse / fc) ** 0.5
        print('*-'*10)
        print('NRMSE = ' + str(nrmse))
        print('*-' * 10)

    # para_y.append(nrmse)  # NRMSE
    #     return pre_step
        return nrmse



def calculate_narma(mu, n, average_degree, num_community, is_layered):
    nrmse = []
    for k in range(1):
        if is_layered:
            W = generate_networks.make_recurrent_layered_network(n, average_degree, num_community, mu)
        else:
            # W = generate_networks.make_modular_network(n, average_degree, num_community, mu)
            W, point = generate_networks.make_modular_network(n, num_community, mu, 0.1)

        # 设置W_IN
        W_IN = np.random.uniform(-1, 1, size=(n, 2))
        # 调整谱半径
        radius = np.max(np.abs(np.linalg.eigvals(W)))
        # spectral_radius = SPECT_RADIUS

        # W = W * (spectral_radius / radius)
        # radius = np.max(np.abs(np.linalg.eigvals(W)))
        print("u={0},n={1},spectral_radius={2}".format(mu, n, num_community))

        # data, target = make_data_for_narma(trainlen + future)
        # 构建模块化ESN
        esn = LI_ESN_internal(n_inputs=1,
                              n_outputs=1,
                              n_reservoir=n,
                              W=W,
                              W_in=W_IN,
                              noise=0,
                              time_scale=time_scale)
        # 拟合训练集与测试
        y = esn.fit_predict(trainLen)
        nrmse.append(y)
    return nrmse

def savedata(row, col, arr):
    data = openpyxl.load_workbook('Result/Data/data.xlsx')
    sheetnames = data.sheetnames
    table = data[sheetnames[1]]
    table = data.active
    rows = row
    # cols = table.max_column
    cols = col
    for i in arr:
        table.cell(rows, cols).value = i
        rows += 4
    data.save('Result/Data/data.xlsx')

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--average_degree', type=int, default=10)
    parser.add_argument('--num_community', type=float, default=5)
    parser.add_argument('--layered', type=bool, default=False)
    args = parser.parse_args()
    print(args)

    res_list = np.arange(400, 401, 400)
    r_list = np.arange(0.1, 1.1, 0.1)
    n_list = np.array([1,2,5,8,10])
    All_data = []

    row = 143
    col = 2
    start_all_time = time.time()
    for r in r_list:
        row = 143
        for res in res_list:
            # 计时看看跑了多久
            start_time = time.time()

            N_node = 500
            # 泄露积分型ESN中的参数a
            a = 1
            time_scale = np.ones(N_node) * a

            SPECT_RADIUS = r
            print("Radius:", SPECT_RADIUS)
            acc = calculate_narma(mu=0.3,
                                  n=N_node,
                                  average_degree=10,
                                  num_community=1,
                                  is_layered=args.layered)
            end_time = time.time() - start_time
            avg = np.mean(acc)
            acc.append(avg)
            print('-'*10, "  n={0},res={1},runtime={2}  ".format(1,res,end_time,".2f"),'-'*10)
            print(acc)
            print('--' * 20)
            # savedata(row, col, acc)
            All_data.append(acc)
            row += 1
        col += 1
    # narma_mean_list = np.array(narma_mean_list)
    # narma_std_list = np.array(narma_std_list)
    print(res_list)
    print(n_list)
    print(All_data)

    # layered == False 表示这是模块化ESN，将结果保存到ｘｌｓｘ中
    # if args.layered == False:
    #     data = openpyxl.load_workbook('Result/Data/data.xlsx')
    #     sheetnames = data.sheetnames
    #     # print(sheetnames)
    #     table = data[sheetnames[5]]
    #     # print(table)
    #     table = data.active
    #     rows = 1
    #     # cols = table.max_column
    #     cols = 2
    #     # print(cols)
    #     for i in para_nrmse:
    #         table.cell(rows + 1, cols).value = i
    #         rows += 1
    #     data.save('Result/Data/data.xlsx')

    start_all_time = time.time() - start_all_time
    print("All Runtime:", start_all_time)
    print("Finish!")