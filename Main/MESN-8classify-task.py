# Leaky integrator model of Echo State Network
import os
# os.environ["CUDA_VISIBLE_DEVICES"] = "6, 7"
import numpy as np
import time
import torch
import generate_networks
import argparse
from sklearn.preprocessing import normalize
import openpyxl
import send_mail as sm
from sklearn.model_selection import KFold
import random

is_cuda = torch.cuda.is_available()
print("Is_GPU:", is_cuda)

# 划分训练集，验证集，测试集
path = "D:\\Pycharm\\Project\\ML\\Dataset\\DEAP\\data_preprocessed_python\\DEAP\\"

# 五折交叉验证
# Trainset data_train Y Z
with open(path + 'data.npy', 'rb') as fileTrain:
    data_train = np.load(fileTrain)
with open(path + 'label.npy', 'rb') as fileTrainL:
    label_train = np.load(fileTrainL)
data = normalize(data_train) # L2范式归一化
# data_train = data_train[:160000, :]
data = torch.tensor(data, dtype=torch.double).cuda()
label = label_train[:, :3]
label = torch.tensor(label, dtype=torch.double).cuda()

index = [i for i in range(len(data))]
seed=42
random.seed(seed)
random.shuffle(index)
data = data[index]
label = label[index]
print("Last index:", index[-1])
print("Data OK!")


'''
# Trainset data_train
with open(path + 'data_train.npy', 'rb') as fileTrain:
    data_train = np.load(fileTrain)
with open(path + 'label_train.npy', 'rb') as fileTrainL:
    label_train = np.load(fileTrainL)
data_train = normalize(data_train) # L2范式归一化
# data_train = data_train[:160000, :]
data_train = torch.tensor(data_train, dtype=torch.double).cuda()
label_train = label_train[:, :3]
label_train = torch.tensor(label_train, dtype=torch.double).cuda()

# Testset
with open(path + 'data_test.npy', 'rb') as fileTest:
    data_test = np.load(fileTest)
with open(path + 'label_test.npy', 'rb') as fileTestL:
    label_test = np.load(fileTestL)

data_test = normalize(data_test)
# data_test = data_test[:40000, :]
data_test = torch.tensor(data_test, dtype=torch.double).cuda()

label_test = label_test[:, :3]
label_test = torch.tensor(label_test, dtype=torch.double).cuda()

print('Train and Test dataset OK!')
# print('-'*20)
len_train = len(data_train)
len_test = len(data_test)

'''


def correct_dimensions(s, targetlength):
    if s is not None:
        s = np.array(s)
        if s.ndim == 0:
            s = np.array([s] * targetlength)
        elif s.ndim == 1:
            if not len(s) == targetlength:
                raise ValueError("arg must have length " + str(targetlength))
        else:
            raise ValueError("Invalid argument")
    return s


def identity(x):
    return x


def step_function(x):
    if x > 0.5:
        return 1
    else:
        return 0


def sigmoid(x):
    return 1 / (1 + np.exp(-10 * x + 1))



class LI_ESN_internal:

    def __init__(self, n_inputs, n_outputs, n_reservoir=500, W=None, W_in=None,
                 noise=0.001, input_shift=None,
                 input_scaling=None, feedback_scaling=None,
                 teacher_scaling=None, teacher_shift=None,
                 out_activation=identity, inverse_out_activation=identity,
                 random_state=None, time_scale=None):
        # check for proper dimensionality of all arguments and write them down.
        self.n_inputs = n_inputs
        self.n_reservoir = n_reservoir
        self.n_outputs = n_outputs
        self.noise = noise
        self.input_shift = correct_dimensions(input_shift, n_inputs)
        self.input_scaling = correct_dimensions(input_scaling, n_inputs)

        self.teacher_scaling = teacher_scaling
        self.teacher_shift = teacher_shift

        self.out_activation = out_activation
        self.inverse_out_activation = inverse_out_activation
        self.random_state = random_state
        self.time_scale = time_scale
        self.W = W
        self.W_in = W_in.to(torch.double)
        # self.W_in = self.W_in.to(torch.double)

        # the given random_state might be either an actual RandomState object,
        # a seed or None (in which case we use numpy's builtin RandomState)
        if isinstance(random_state, np.random.RandomState):
            self.random_state_ = random_state
        elif random_state:
            try:
                self.random_state_ = np.random.RandomState(random_state)
            except TypeError as e:
                raise Exception("Invalid seed: " + str(e))
        else:
            self.random_state_ = np.random.mtrand._rand

    def _scale_inputs(self, inputs):
        """for each input dimension j: multiplies by the j'th entry in the
        input_scaling argument, then adds the j'th entry of the input_shift
        argument."""
        if self.input_scaling is not None:
            inputs = np.dot(inputs, np.diag(self.input_scaling))
        if self.input_shift is not None:
            inputs = inputs + self.input_shift
        return inputs

    def _update(self, state, input_pattern):
        # leaky integrator model:
        # it can adjust timescales for each neurons.
        # print(self.W.dtype, state.dtype, self.W_in.dtype, input_pattern.dtype)
        # print(state.shape, input_pattern.shape)
        if type(state) == np.ndarray:
            state = torch.tensor(state)
        state = state.unsqueeze(1)
        input_pattern = input_pattern.unsqueeze(1)
        state = state.to(torch.double)

        # print(self.W.dtype, state.dtype, self.W_in.dtype, input_pattern.dtype)
        self.W = self.W.cuda()
        self.W_in = self.W_in.cuda()
        # print(self.W.is_cuda, state.is_cuda, self.W_in.is_cuda, input_pattern.is_cuda)

        preactivation = (torch.mm(self.W, state) + torch.mm(self.W_in, input_pattern)).cuda()

        # print(self.time_scale.is_cuda, state.is_cuda)
        self.time_scale = self.time_scale.cuda()
        state = ((1 - self.time_scale) * state + self.time_scale * torch.tanh(preactivation)).cuda()
        # state = (1 - self.time_scale) * state + self.time_scale * sigmoid(preactivation)

        # t1 = (torch.rand(self.n_reservoir) - 0.5) * self.noise * self.time_scale
        t2 = state

        return t2

    def calc_lyapunov_exp(self, inputs, initial_distance, n):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        states1 = np.zeros((inputs.shape[0], self.n_reservoir))
        states2 = np.zeros((inputs.shape[0], self.n_reservoir))
        transient = min(int(inputs.shape[0] / 10), 100)
        for i in range(1, transient):
            states1[i, :] = self._update(states1[i - 1], inputs[i, :])
        states2[transient - 1, :] = states1[transient - 1, :]
        states2[transient - 1, n] = states2[transient - 1, n] + initial_distance
        gamma_k_list = []
        for k in range(transient, inputs.shape[0]):
            states1[k, :] = self._update(states1[k - 1], inputs[k, :])
            states2[k, :] = self._update(states2[k - 1], inputs[k, :])
            gamma_k = np.linalg.norm(states2[k, :] - states1[k, :])
            gamma_k_list.append(gamma_k / initial_distance)
            states2[k, :] = states1[k, :] + (initial_distance / gamma_k) * (states2[k, :] - states1[k, :])
        lyapunov_exp = np.mean(np.log(gamma_k_list))
        return lyapunov_exp

    def fit(self, inputs, outputs):
        if inputs.ndim < 2:
            inputs = inputs.resize_(len(inputs), 1)
        if outputs.ndim < 2:
            outputs = outputs.resize_(len(outputs), 1)
        inputs_scaled = self._scale_inputs(inputs)
        teachers_scaled = outputs


        # step the reservoir through the given input,output pairs:
        # states = np.zeros((inputs.shape[0], self.n_reservoir))
        # 设置状态为均值为0，方差为0.1的正态分布，使参数更容易收敛。
        with torch.no_grad():
            states = torch.normal(0, 0.1**0.5, (inputs.shape[0], self.n_reservoir)).cuda()
        # print(type(inputs))
        # print(inputs.shape[0])
        for n in range(1, inputs.shape[0]):
            tem = self._update(states[n - 1], inputs_scaled[n, :])
            # print(type(tem))
            # print(type(states[n, :]))
            # print(len(states[n, :]))
            # states[n, :] = tem
            states[n, :] = tem[:,0]
        transient = min(int(inputs.shape[0] / 10), 100)
        # print(type(states))
        # print(type(states))
        # print(type(inputs_scaled))
        # states = states.cpu().numpy()
        extended_states = torch.cat((states, inputs), dim =1)
        # extended_states = np.hstack((states, inputs_scaled))
        pinv = torch.linalg.pinv(extended_states[transient:, :])
        teachers_scaled = teachers_scaled.cuda()
        self.W_out = torch.mm(pinv, teachers_scaled[transient:, :]).T
        # self.W_out = np.dot(np.linalg.pinv(extended_states[transient:, :]), teachers_scaled[transient:, :]).T
        # print(self.W_out.shape)
        # print(type(self.W_out))
        self.W_out = torch.as_tensor(self.W_out)
        # remember the last state for later:
        self.laststate = states[-1, :]
        self.lastinput = inputs[-1, :]
        self.lastoutput = teachers_scaled[-1, :]

        # apply learned weights to the collected states:
        pred_train = torch.mm(extended_states, self.W_out.T)
        # pred_train = np.dot(extended_states, self.W_out.T)
        return pred_train

    def predict(self, inputs, continuation=True):
        if inputs.ndim < 2:
            inputs = np.reshape(inputs, (len(inputs), -1))
        n_samples = inputs.shape[0]

        if continuation:
            laststate = self.laststate
            lastinput = self.lastinput
            lastoutput = self.lastoutput
        else:
            laststate = np.zeros(self.n_reservoir)
            lastinput = np.zeros(self.n_inputs)
            lastoutput = np.zeros(self.n_outputs)

        inputs = torch.as_tensor(inputs)
        # laststate = laststate.T
        # print(laststate.is_cuda)
        laststate = torch.unsqueeze(laststate, dim=1)
        laststate = laststate.T
        states = torch.cat((laststate, torch.zeros((n_samples, self.n_reservoir)).cuda()), dim=0)
        # states = np.vstack(
        #     [laststate, np.zeros((n_samples, self.n_reservoir))])

        lastoutput = torch.unsqueeze(lastoutput, dim=1)
        lastoutput = lastoutput.T
        outputs = torch.cat((lastoutput, torch.zeros((n_samples, self.n_outputs)).cuda()), dim=0)
        # outputs = np.vstack(
        #     [lastoutput, np.zeros((n_samples, self.n_outputs))])

        # print(inputs.shape) # (78080, 128)
        # print(states.shape) # (78081, 200)
        states = torch.as_tensor(states)
        outputs = torch.as_tensor(outputs)
        for n in range(n_samples):
            tem = self._update(states[n, :], inputs[n, :])
            states[n + 1, :] = tem[:,0]
            # outputs[n + 1, :] = np.dot(self.W_out, np.concatenate([states[n + 1, :], inputs[n, :]]))
            t1 = torch.cat((states[n+1, :], inputs[n,:]), dim =0)
            # print(t1.shape)
            t1 = torch.unsqueeze(t1, dim=1)
            wt1 = torch.mm(self.W_out, t1)
            wt1 = wt1.T
            outputs[n+1: ] = wt1

        # print("outputs.shape:", outputs.shape)
        return self.out_activation(outputs[1:])
        # print(outputs[1:])
        # return np.heaviside(outputs[1:]-0.5, 0)*0.3

def emotion(e0, e1, e2):
    if e0 > 5 and e1 <= 5 and e2 <= 5:  # Protected
        return 1
    elif e0 > 5 and e1 <= 5 and e2 > 5:  # Satisfied
        return 2
    elif e0 > 5 and e1 > 5 and e2 <= 5:  # Surprised
        return 3
    elif e0 > 5 and e1 > 5 and e2 > 5:  # Happy
        return 4
    elif e0 <= 5 and e1 <= 5 and e2 <= 5:  # Sad
        return 5
    elif e0 <= 5 and e1 <= 5 and e2 > 5:  # Unconcerned
        return 6
    elif e0 <= 5 and e1 > 5 and e2 <= 5:  # Frightened
        return 7
    elif e0 <= 5 and e1 > 5 and e2 > 5:  # Angry
        return 8

def compute_acc(jz):
    acc, s = 0, 0
    for i in range(1, len(jz)):
        acc += jz[i][i]
        s += sum(jz[i])
    return acc / s

def calculate_narma(p1, p2, n, average_degree, m, is_layered=False):

    for k in range(1):
        if is_layered == True:
            W = generate_networks.make_recurrent_layered_network(n, average_degree, m, p1)
        elif is_layered == False and m != 1:
            W, point = generate_networks.make_modular_network(n, average_degree, m, p1, p2)
        elif m == 1:
            W = generate_networks.make_ESN(n, p1)

        W = torch.tensor(W, dtype=torch.double)
        # 设置W_IN
        W_IN = torch.Tensor(n, 128).uniform_(-0.1, 0.1)

        # 调整谱半径
        # radius = np.max(np.abs(np.linalg.eigvals(W)))
        rad = torch.linalg.eigvals(W)
        # print(rad)
        # print(type(rad))
        radius = torch.max(torch.abs(rad))
        # TypeError: abs(): argument 'input' (position 1) must be Tensor, not torch.return_types.eig

        W = W * (SPECT_RADIUS / radius)
        # rad = torch.eig(W, eigenvectors=False).eigenvalues
        rad = torch.linalg.eigvals(W)
        radius = torch.max(torch.abs(rad))
        print("spectral_radius:", radius)

        # data, target = make_data_for_narma(trainlen + future)
        # 构建模块化ESN
        esn = LI_ESN_internal(n_inputs=128,
                              n_outputs=3,
                              n_reservoir=n,
                              W=W,
                              W_in=W_IN,
                              noise=0.00,
                              time_scale=time_scale)
        # 拟合训练集
        with torch.no_grad():
            esn.fit(data_train, label_train).cuda()
        print("Fitting OK!")
        # 预测测试集
        with torch.no_grad():
            prediction = esn.predict(data_test).cuda()
        # print(type(prediction))
        prediction = torch.as_tensor(prediction)
        print("Predicting OK!")
        result = np.zeros(shape=(9,9))
        t, all_num = 0, 0
        # print(type(L))
        for i in range(len_test):
            true_label = emotion(label_test[i][0], label_test[i][1], label_test[i][2])
            all_num += 1
            pre_label = emotion(prediction[i][0], prediction[i][1], prediction[i][2])
            result[true_label][pre_label] += 1

        acc = compute_acc(result)
        # print("Classification accuracy:", t, all_num, acc*100)

        # narma_result = np.sqrt(
        #     np.mean((np.reshape(prediction, -1) - np.reshape(target[2200:], -1)) ** 2) / np.var(target[2200:]))
        # narma_list.append(narma_result)
    return acc


def save(row, col, result):
    xlsx_path = 'Result/AAAI Data/test.xlsx'  # xlsx路径
    data = openpyxl.load_workbook(xlsx_path)
    sheetnames = data.sheetnames
    table = data[sheetnames[0]] # 指定xlsx表中的sheet
    table = data.active
    table.cell(row, col).value = result
    data.save(xlsx_path)

if __name__ == '__main__':
    all_time = time.time()

    # 结果
    Result = []
    Acc_max = 1.5

    # 自变量
    m_list = np.array([1, 2, 3, 4, 5, 6, 8, 10])  # 模块化数量 Modularity
    # m_list = np.array([1])
    p1_list = np.arange(0.20, 0.201, 0.05)  # 模块内连接概率
    p2_list = np.arange(0.01, 0.051, 0.01)  # 模块间连接概率
    n_list = np.array([1440, 2160])  # 储层节点数
    r_list = np.arange(0.8, 1.01, 0.05)  # 谱半径
    l_list = np.arange(0.06, 0.061, 0.01)  # 泄露率

    row, col = 4, 2  # 保存所有指标

    # kf = KFold(n_splits=5, shuffle=False)  # 初始化KFold 5折
    for m in m_list:
        for p2 in p2_list:
            # res = 0
            tem = []
            start_time = time.time()  # 计时单次运行时间

            N_node = 1200  # 储层节点数
            a = 1  # 泄露积分型中的参数a
            time_scale = torch.ones(N_node, 1) * a
            SPECT_RADIUS = 0.85  # 谱半径
            # Result Degree
            result = calculate_narma(p1=0.05,
                                     p2=p2,
                                     n=N_node,
                                     average_degree=10,
                                     m=m,
                                     is_layered=False,
                                     )
            end_time = time.time() - start_time
            # res += result
            print('***' * 24)
            print('---' * 11 + "Result" + '---' * 11)
            print("p1={0}, p2={1}, n={2}, m={3}, r={4}, a={5}, runtime={6},"
                  .format(0.05, round(p2, 2), N_node, m, SPECT_RADIUS, a, round(end_time, 2)))
            # print("result = {0}".format(result))
            print("Acc={0}".format(result))
            print('---' * 24)
            print('***' * 24)

            # tem.extend([result])
            save(row, col, result)
            Result.append(result)
        # print("五折交叉验证后结果为：",res/5)
        row += 1

        col += 1
        row = 4

    # print("p1_list:", p1_list)
    print("p2_list:", p2_list)
    print("m_list:", m_list)
    print("Result:", Result)

    # 发送邮件
    all_time = time.time() - all_time
    print("工作站 All Runtime:{0}s".format(all_time))
    context = "工作站 Runtime:" + str(all_time) + "s \nOK!"
    sm.send_mail(context)

